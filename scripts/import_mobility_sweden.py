import argparse
import csv
import json
import re
import unicodedata
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from urllib.parse import unquote, urljoin
from urllib.request import Request, urlopen

from openpyxl import load_workbook


BASE_URL = "https://mobilitysweden.se"
LATEST_PAGE = (
    "https://mobilitysweden.se/statistik/Nyregistreringar_per_manad_1/"
    "nyregistreringar-2026/stark-marsmanad-lyfter-kvartalet-elfordon-okar-i-samtliga-segment"
)

MONTHS = {
    "januari": 1,
    "februari": 2,
    "mars": 3,
    "april": 4,
    "maj": 5,
    "juni": 6,
    "juli": 7,
    "augusti": 8,
    "september": 9,
    "oktober": 10,
    "november": 11,
    "december": 12,
}

BRAND_ALIASES = {
    "VW": "VOLKSWAGEN",
    "MERCEDES": "MERCEDES-BENZ",
    "MERCEDES BENZ": "MERCEDES-BENZ",
}

BRAND_PREFIXES = [
    "LYNK & CO",
    "MERCEDES BENZ",
    "MERCEDES",
    "VOLVO",
    "TESLA",
    "VW",
    "VOLKSWAGEN",
    "POLESTAR",
    "KIA",
    "TOYOTA",
    "SKODA",
    "BMW",
    "RENAULT",
    "AUDI",
    "FORD",
    "CUPRA",
    "ZEEKR",
    "MAZDA",
    "MG",
    "LEXUS",
    "XPENG",
    "SUBARU",
    "PORSCHE",
    "MINI",
    "HYUNDAI",
    "OPEL",
    "NISSAN",
    "PEUGEOT",
    "CITROEN",
    "BYD",
    "FIAT",
    "SMART",
    "HONDA",
    "JEEP",
]


@dataclass
class MarketModel:
    brand_raw: str
    model_raw: str
    fuel_type_raw: str
    normalized_brand: str
    normalized_model: str
    model_group: str | None
    needs_mapping: bool
    first_seen_month: str
    last_seen_month: str
    registrations_last_month: int
    registrations_ytd: int
    registrations_12m: int | None
    source_name: str
    source_url: str


@dataclass
class MonthlyStat:
    market_model_key: str
    month: str
    brand_raw: str
    model_raw: str
    fuel_type_raw: str
    registrations: int


def fetch_text(url: str) -> str:
    request = Request(url, headers={"User-Agent": "swedish-ev-advisor/0.1"})
    with urlopen(request, timeout=30) as response:
        return response.read().decode("utf-8", errors="replace")


def download(url: str, target: Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    request = Request(url, headers={"User-Agent": "swedish-ev-advisor/0.1"})
    with urlopen(request, timeout=60) as response:
        target.write_bytes(response.read())


def find_monthly_report_url(page_url: str) -> str:
    html = fetch_text(page_url)
    matches = re.findall(r'href="([^"]+\.xlsx[^"]*)"', html, flags=re.I)
    for href in matches:
        decoded = unquote(href)
        if "Månadsrapport Nyregistreringar" in decoded:
            return urljoin(BASE_URL, href)
    raise RuntimeError(f"No Mobility Sweden monthly report .xlsx found on {page_url}")


def parse_month_from_workbook_title(title: str) -> date:
    match = re.search(r"Nyregistreringar\s+([A-Za-zÅÄÖåäö]+)\s+(\d{4})", title or "")
    if not match:
        raise RuntimeError(f"Could not parse report month from title: {title!r}")
    month_name = match.group(1).lower()
    return date(int(match.group(2)), MONTHS[month_name], 1)


def normalize(value: str) -> str:
    value = unicodedata.normalize("NFKD", value.upper())
    value = "".join(char for char in value if not unicodedata.combining(char))
    value = value.replace("&", " AND ")
    value = re.sub(r"[^A-Z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip().lower()


def split_brand_model(raw_name: str) -> tuple[str, str]:
    name = re.sub(r"\s+", " ", raw_name.strip().upper())
    if name.startswith("MAZDA6E"):
        return "MAZDA", "6E"
    for brand in sorted(BRAND_PREFIXES, key=len, reverse=True):
        if name == brand or name.startswith(brand + " "):
            model = name[len(brand) :].strip()
            return brand, model or "UNKNOWN"
    parts = name.split(" ", 1)
    return parts[0], parts[1] if len(parts) > 1 else "UNKNOWN"


def canonical_brand(raw_brand: str) -> str:
    return BRAND_ALIASES.get(raw_brand, raw_brand)


def is_ambiguous(model_raw: str) -> bool:
    return "/" in model_raw or "ÖVRIGA" in model_raw or "OVRIGA" in normalize(model_raw).upper()


def parse_elbil_ranking(xlsx_path: Path, source_url: str) -> tuple[list[MarketModel], list[MonthlyStat]]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook["Elbil ranking"]
    report_month = parse_month_from_workbook_title(worksheet.cell(row=2, column=8).value)
    month_iso = report_month.isoformat()
    market_models: list[MarketModel] = []
    monthly_stats: list[MonthlyStat] = []

    for row in worksheet.iter_rows(min_row=10, values_only=True):
        full_name = row[3]
        registrations = row[5]
        ytd = row[12]
        if not full_name or str(full_name).strip().upper() == "TOTALT":
            continue
        if not isinstance(registrations, int) or not isinstance(ytd, int):
            continue

        brand_raw, model_raw = split_brand_model(str(full_name))
        normalized_brand = normalize(canonical_brand(brand_raw))
        normalized_model = normalize(model_raw)
        needs_mapping = is_ambiguous(model_raw)
        model_group = str(full_name).strip().upper() if needs_mapping else None
        key = f"{normalized_brand}:{normalized_model}:EL"

        market_models.append(
            MarketModel(
                brand_raw=brand_raw,
                model_raw=model_raw,
                fuel_type_raw="EL",
                normalized_brand=normalized_brand,
                normalized_model=normalized_model,
                model_group=model_group,
                needs_mapping=needs_mapping,
                first_seen_month=month_iso,
                last_seen_month=month_iso,
                registrations_last_month=registrations,
                registrations_ytd=ytd,
                registrations_12m=None,
                source_name="Mobility Sweden",
                source_url=source_url,
            )
        )
        monthly_stats.append(
            MonthlyStat(
                market_model_key=key,
                month=month_iso,
                brand_raw=brand_raw,
                model_raw=model_raw,
                fuel_type_raw="EL",
                registrations=registrations,
            )
        )

    return market_models, monthly_stats


def write_csv(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        return
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--page-url", default=LATEST_PAGE)
    parser.add_argument("--xlsx", type=Path)
    parser.add_argument("--out-dir", type=Path, default=Path("data/mobility-sweden/processed"))
    parser.add_argument("--raw-dir", type=Path, default=Path("data/mobility-sweden/raw"))
    args = parser.parse_args()

    source_url = args.page_url
    if args.xlsx:
        xlsx_path = args.xlsx
    else:
        xlsx_url = find_monthly_report_url(args.page_url)
        filename = unquote(xlsx_url.rstrip("/").split("/")[-1])
        xlsx_path = args.raw_dir / filename
        download(xlsx_url, xlsx_path)
        source_url = xlsx_url

    market_models, monthly_stats = parse_elbil_ranking(xlsx_path, source_url)
    args.out_dir.mkdir(parents=True, exist_ok=True)

    market_rows = [asdict(row) for row in market_models]
    monthly_rows = [asdict(row) for row in monthly_stats]
    write_csv(args.out_dir / "market_models.csv", market_rows)
    write_csv(args.out_dir / "market_model_monthly_stats.csv", monthly_rows)
    (args.out_dir / "market_models.json").write_text(
        json.dumps(market_rows, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (args.out_dir / "market_model_monthly_stats.json").write_text(
        json.dumps(monthly_rows, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    needs_mapping = [row for row in market_models if row.needs_mapping]
    print(
        f"Imported {len(market_models)} electric passenger-car model rows "
        f"from {xlsx_path.name}. {len(needs_mapping)} need manual alias mapping."
    )
    if needs_mapping:
        print("Needs mapping:", ", ".join(row.model_group or row.model_raw for row in needs_mapping[:12]))


if __name__ == "__main__":
    main()
